# ==============================
# FastAPI Excel Template Report API
# Para deploy no Render.com
# Recebe JSON do Lovable
# Preenche template Excel fixo
# Retorna arquivo Excel pronto
# ==============================

# -------- requirements.txt --------
# fastapi
# uvicorn
# openpyxl
# pydantic


# -------- main.py --------

from fastapi import FastAPI
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from openpyxl import load_workbook
import io

app = FastAPI(title="Template Report API")

TEMPLATE_PATH = "RELATORIO_TEMPLATE.xlsx"  # coloque o template no repo


# -------------------------------
# Modelo de dados recebido do Lovable
# -------------------------------

class ReportData(BaseModel):
    recebimento: str
    inicio: str
    finalizado: str
    emissao: str

    tipoprojeto: str
    quantidade: str
    modelo: str
    capnominalAH: str
    tensaonominalV: str

    descricao: str
    objetivo: str
    metodo: str

    equipamento: str
    fabricante: str
    modelo_equip: str
    identificacao: str

    local: str
    temperatura: str
    umidade: str


@app.get("/")
def root():
    return {"status": "ok", "service": "template-report-api"}


# -------------------------------
# Util — escrever valor em range mesclado
# -------------------------------

def write_range(ws, cell_range, value):
    start = cell_range.split(":")[0]
    ws[start] = value


# -------------------------------
# Endpoint — gerar relatório
# -------------------------------

@app.post("/generate-report")
def generate_report(data: ReportData):

    wb = load_workbook(TEMPLATE_PATH)
    ws = wb.active

    # ====== CAMPOS CONFORME COORDENADAS FORNECIDAS ======

    write_range(ws, "K2:L2", data.recebimento)
    write_range(ws, "K3:L3", data.inicio)
    write_range(ws, "K4:L4", data.finalizado)
    write_range(ws, "K5:L5", data.emissao)

    write_range(ws, "C14:K14", data.tipoprojeto)

    write_range(ws, "B19:C19", data.quantidade)
    write_range(ws, "D19:F19", data.modelo)
    write_range(ws, "G19:H19", data.capnominalAH)
    write_range(ws, "I19:J19", data.tensaonominalV)

    write_range(ws, "B22:J24", data.descricao)
    write_range(ws, "B27:K29", data.objetivo)
    write_range(ws, "B33:I44", data.metodo)

    write_range(ws, "B47:C57", data.equipamento)
    write_range(ws, "D47:E57", data.fabricante)
    write_range(ws, "F47:G57", data.modelo_equip)
    write_range(ws, "H47:I57", data.identificacao)

    ws["E62"] = data.local
    ws["E63"] = data.temperatura
    ws["E64"] = data.umidade

    # ====== GERAR ARQUIVO DE SAÍDA ======

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=relatorio_preenchido.xlsx"
        }
    )


# -------------------------------
# Deploy Render
# -------------------------------
# Build command:
# pip install -r requirements.txt
#
# Start command:
# uvicorn main:app --host 0.0.0.0 --port 10000
